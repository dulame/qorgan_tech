"""
Excel Exporter for Password Check Data
Exports password check logs from SQLite database to Excel format
"""

import os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from modules.database import Database


class PasswordCheckExporter:
    """Export password check data to Excel"""

    def __init__(self, output_dir='data'):
        """
        Initialize the exporter.
        
        Args:
            output_dir (str): Directory to save Excel files
        """
        self.output_dir = output_dir
        Path(output_dir).mkdir(exist_ok=True)
        self.db = Database()

    def export_to_excel(self, filename='password_checks.xlsx'):
        """
        Export all password checks to Excel.
        
        Args:
            filename (str): Name of the Excel file. Defaults to 'password_checks.xlsx'
                           (this file will be overwritten on each export)
            
        Returns:
            str: Path to the created Excel file
        """
        filepath = os.path.join(self.output_dir, filename)
        
        # Get data from database
        checks = self.db.get_all_password_checks()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Password Checks"
        
        # Add headers
        headers = [
            'User ID',
            'Check Timestamp',
            'Hashed Password',
            'Strength Score',
            'Strength Level',
            'Time to Crack (Offline)',
            'Crack Time (Seconds)'
        ]
        
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for check in checks:
            row = [
                check['user_id'],
                check['checked_at'],
                check['hashed_password'],
                check['strength_score'],
                check['strength_text'],
                check['crack_time_offline'],
                check['crack_time_seconds']
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        column_widths = {
            'A': 12,    # User ID
            'B': 20,    # Check Timestamp
            'C': 50,    # Hashed Password
            'D': 16,    # Strength Score
            'E': 18,    # Strength Level
            'F': 25,    # Time to Crack
            'G': 18     # Crack Time Seconds
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Data rows
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Add summary sheet
        self._add_summary_sheet(wb, checks)
        
        # Save workbook
        wb.save(filepath)
        
        return filepath

    def _add_summary_sheet(self, workbook, checks):
        """Add a summary sheet with statistics"""
        ws = workbook.create_sheet("Summary")
        
        # Calculate statistics
        total_checks = len(checks)
        unique_users = len(set(check['user_id'] for check in checks))
        
        strength_distribution = {
            'Very Weak': 0,
            'Weak': 0,
            'Fair': 0,
            'Good': 0,
            'Very Strong': 0
        }
        
        for check in checks:
            strength_text = check['strength_text']
            if strength_text:
                for key in strength_distribution:
                    if key.lower() in strength_text.lower():
                        strength_distribution[key] += 1
                        break
        
        # Add summary data
        ws.append(['Password Check Summary Report'])
        ws.append([])
        ws.append(['Total Checks', total_checks])
        ws.append(['Unique Users', unique_users])
        ws.append([])
        ws.append(['Strength Distribution'])
        
        for strength, count in strength_distribution.items():
            ws.append([strength, count])
        
        # Style summary sheet
        ws['A1'].font = Font(bold=True, size=14)
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        
        return ws

    def export_user_checks(self, user_id, filename=None):
        """
        Export password checks for a specific user.
        
        Args:
            user_id (int): The user ID to filter
            filename (str): Name of the Excel file. If None, uses default name
                           (this file will be overwritten on each export for the user)
            
        Returns:
            str: Path to the created Excel file
        """
        if filename is None:
            filename = f"password_checks_user_{user_id}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # Get data from database
        checks = self.db.get_password_checks(user_id)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"User {user_id}"
        
        # Add headers
        headers = [
            'Check Timestamp',
            'Hashed Password',
            'Strength Score',
            'Strength Level',
            'Time to Crack (Offline)',
            'Crack Time (Seconds)'
        ]
        
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for check in checks:
            row = [
                check['checked_at'],
                check['hashed_password'],
                check['strength_score'],
                check['strength_text'],
                check['crack_time_offline'],
                check['crack_time_seconds']
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        column_widths = {
            'A': 20,    # Check Timestamp
            'B': 50,    # Hashed Password
            'C': 16,    # Strength Score
            'D': 18,    # Strength Level
            'E': 25,    # Time to Crack
            'F': 18     # Crack Time Seconds
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save workbook
        wb.save(filepath)
        
        return filepath


def main():
    """Main function for standalone usage"""
    exporter = PasswordCheckExporter()
    filepath = exporter.export_to_excel()
    print(f"✅ Password checks exported to: {filepath}")


if __name__ == "__main__":
    main()
